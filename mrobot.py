import json, sys, os
from openpyxl import Workbook, load_workbook
from MomondoMiner import MomondoMiner

def mine_and_save_to_file(output_file, m):
    if not m.initiate_session():
        print "Error initizling session"
        sys.exit(1)
    if not m.get_search_id():
        print "Error while searching"
        sys.exit(1)

    success = m.mine()
    if not success:
        sys.exit(0)

    data = m.get_best_offer()

    print json.dumps(data, indent=4, sort_keys=True)
    print "\nGenerating workbook file..."


    if os.path.isfile(output_file) and os.path.exists(output_file):
        wb = load_workbook(filename = output_file)
    else:
        wb = Workbook()
    ws = wb.active

    ws.title = "MomondoMiner"

    ws['A1'] = "Airline"
    ws['B1'] = "Origin"
    ws['C1'] = "Destination"
    ws['D1'] = "Departure Date"
    ws['E1'] = "Return Date"
    ws['F1'] = "Total Price"
    ws['G1'] = "Currency"
    ws['H1'] = "Departure Duration"
    ws['I1'] = "Return Duration"
    ws['J1'] = "Departure Stops"
    ws['K1'] = "Return Stops"
    ws['L1'] = "URL"

    ws.append({'A': data["journey_segments"][0]["segment_legs"][0]["airline_name"],
        'B': data["origin"],
        'C' : data["destination"],
        'D' : data["departure"],
        'E' : data["arrival"],
        'F' : data["total_price"],
        'G' : data["currency_type"],
        'H' : data["journey_segments"][0]["segment_duration"],
        'I' : data["journey_segments"][1]["segment_duration"],
        'J' : data["journey_segments"][0]["segment_stops"],
        'K' : data["journey_segments"][1]["segment_stops"],
        'L' : data["offer_url"]})

    wb.save(output_file)

    print "Done!"

if __name__ == "__main__":
    output_file = "MomondoMiner.xlsx"
    mine_and_save_to_file(output_file, MomondoMiner("SFO", "LAX", "2017-12-24", "2017-12-27"))
